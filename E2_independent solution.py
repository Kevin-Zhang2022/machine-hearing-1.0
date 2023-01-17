import numpy as np
import snntorch as snn
import torch
import numpy as Num
import math as Mat
import torch.nn as nn
import matplotlib.pyplot as Plt
from snntorch import spikeplot as splt
from snntorch import spikegen
import openpyxl
import os
import torch
# from E0_Net import Net
import torch.nn as nn

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse, Circle
from torch.utils.data import TensorDataset, DataLoader

from matplotlib import pyplot as plt
from DatPro import DatPro


def seed_torch(seed=1):
    np.random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.backends.cudnn.benchmark = False
    torch.backends.cudnn.deterministic = True
seed_torch()
# s_1000 s_2000 s_3000 d_1000-3000
# DatSet = DatPro(Sec=10, Rat_Tra=0.8, keywords=keywords)
#
# Tra = []
# Tes = []
# batch_size = 128
# for i in range(6):
#     Tra.append(DataLoader(DatSet[i*2], batch_size=batch_size, shuffle=True, drop_last=True))
#     Tes.append(DataLoader(DatSet[i*2+1], batch_size=batch_size, shuffle=True, drop_last=True))

################################################
ind_tra = 1
keywords = 's_1000'

for i in range(6):
    StrCom = f"Net_{i:d}"+" = torch.load('Net/E0/"+keywords+f"/Net_{i}_{ind_tra:d}.pkl')"
    exec(StrCom)

for i in range(3):
    for j in range(2):
        StrCom = f'W_{i:d}_{j:d}_0=Net_{i:d}.fc{j+1}_0.weight.detach().numpy()'
        exec(StrCom)

for i in range(3,6):
    StrCom = f'W_{i:d}_0_0=Net_{i:d}.fc1_0.weight.detach().numpy()'
    exec(StrCom)
    StrCom = f'W_{i:d}_0_1=Net_{i:d}.fc1_1.weight.detach().numpy()'
    exec(StrCom)
    StrCom = f'W_{i:d}_1_0=Net_{i:d}.fc2_0.weight.detach().numpy()'
    exec(StrCom)




#### net 5
fig0 = plt.figure(figsize=(9,6))
fontsize = 8

axe = fig0.add_subplot(2, 1, 1)
a = np.average(np.transpose(abs(W_5_1_0)),axis=1)  # 10 60 84 138
axe.plot(a)
axe.set_title('AAW_5_1_0')
axe.set_xlabel('Neuron index',fontsize=fontsize)
axe.set_ylabel('AAW',fontsize=fontsize)

a = Num.transpose(W_5_0_0)[:,10]
b = Num.transpose(W_5_0_0)[:,60]
c = Num.transpose(W_5_0_0)[:,84]
d = Num.transpose(W_5_0_1)[:,138-100]

axe = fig0.add_subplot(2, 2, 3)
axe.plot(b,label='Neuron 60')
axe.plot(c,label='Neuron 84')
axe.plot(d,label='Neuron 138')
axe.legend(loc= 'upper right', fontsize=8)
axe.set_title('Pre-weights of neurons in pattern 0',fontsize=fontsize)
axe.set_ylim([-3,6])
axe.set_xlabel('Neuron index',fontsize=fontsize)
axe.set_ylabel('Pre-weights',fontsize=fontsize)

axe = fig0.add_subplot(2, 2, 4)
axe.plot(b,label='Neuron 60',color='#1f77b4')
axe.plot(a,label='Neuron 10',color='#df2728')
axe.set_ylim([-3,6])
axe.legend(loc='upper right', fontsize=8)
axe.set_title('Pre-weights of neurons in pattern 1 ', fontsize=fontsize)
axe.set_xlabel('Neuron index',fontsize=fontsize)
axe.set_ylabel('Pre-weights',fontsize=fontsize)

fig0.subplots_adjust(wspace=0.221,hspace=0.355,top=0.929,right=0.948,left=0.083,bottom=0.102)
filename = 'Fig/E2/Net5 pattern.pdf'
fig0.savefig(filename)


## corvariance coefficients
ar = np.reshape(a,(300,1))
br = np.reshape(b,(300,1))
cr = np.reshape(c,(300,1))
dr = np.reshape(d,(300,1))

df = np.concatenate((ar,br,cr,dr),axis=1)
corcoe = np.corrcoef(df,rowvar=False)

# wb = openpyxl.Workbook()
# ws_corcoe = wb.create_sheet('corcoe')
# for i in range(corcoe.shape[0]):  # Los.shape[0]
#     for j in range(corcoe.shape[1]):  # Los.shape[1]
#         ws_corcoe.cell(row=i+1, column=j+1, value=corcoe[i][j])
# wb.save('Table/E2/Net5PatternCorCoe.xls')  ## table e0

### net 0
fig1 = plt.figure(figsize=(9,7))
fontsize = 8
axe = fig1.add_subplot(2, 1, 1)
a = np.average(np.transpose(abs(W_0_1_0)),axis=1)  # 6 15 45 51 64 162
axe.set_title('AAW_0_1_0')
axe.set_xlabel('Neuron index',fontsize=fontsize)
axe.set_ylabel('AAW',fontsize=fontsize)
axe.plot(a)

# pattern0
a = Num.transpose(W_0_0_0)[:,6]
b = Num.transpose(W_0_0_0)[:,15]
c = Num.transpose(W_0_0_0)[:,45]
d = Num.transpose(W_0_0_0)[:,51]
e = Num.transpose(W_0_0_0)[:,64]
# pattern1
f = Num.transpose(W_0_0_0)[:,162]

axe = fig1.add_subplot(2, 2, 3)
axe.plot(a, label='Neuron 6')
axe.plot(b, label='Neuron 15')
axe.plot(c, label='Neuron 45')
axe.plot(d, label='Neuron 51')
axe.plot(e, label='Neuron 64')
axe.set_title('Pre-weights of neurons in pattern0', fontsize=fontsize)
axe.set_ylim([-3,6])
axe.legend(loc='upper right', fontsize=8)
axe.set_xlabel('Neuron index',fontsize=fontsize)
axe.set_ylabel('Pre-weights',fontsize=fontsize)


axe = fig1.add_subplot(2, 2, 4)
axe.plot(a, label='Neuron 6')
axe.plot(f, label='Neuron 162',color='#8c564b')
axe.legend(loc= 'upper right', fontsize=8)
axe.set_title('Pre-weights of neurons in pattern1 ', fontsize=fontsize)
axe.set_ylim([-3,6])
axe.set_xlabel('Neuron index',fontsize=fontsize)
axe.set_ylabel('Pre-weights',fontsize=fontsize)

fig1.subplots_adjust(wspace=0.221,hspace=0.355,top=0.929,right=0.948,left=0.083,bottom=0.102)
filename = 'Fig/E2/Net0 pattern.pdf'
fig1.savefig(filename)


#### net0 net 5 similar?
fig2 = plt.figure(figsize=(9,3.5))
fontsize = 8

a = Num.transpose(W_5_0_0)[:,60] # pattern0
b = Num.transpose(W_0_0_0)[:,6]

c = Num.transpose(W_5_0_0)[:,10]
d = Num.transpose(W_0_0_0)[:,162]

axe = fig2.add_subplot(1, 2, 1)
axe.plot(a, label='Pattern0 in net5')
axe.plot(b, label='Pattern0 of net0')
axe.set_title('Pattern0 in net5 and net0', fontsize=fontsize)
axe.legend(loc= 'upper right', fontsize=8)
axe.set_ylim([-3,6])
axe.set_xlabel('Neuron index',fontsize=fontsize)
axe.set_ylabel('Pre-weights',fontsize=fontsize)


axe = fig2.add_subplot(1, 2, 2)
axe.plot(c, label='Patter1 in net5')
axe.plot(d, label='Patter1 in net0')
axe.set_title('Pattern1 in net5 and net0', fontsize=fontsize)
axe.legend(loc='upper right', fontsize=8)
axe.set_ylim([-3,6])
axe.set_xlabel('Neuron index',fontsize=fontsize)
axe.set_ylabel('Pre-weights',fontsize=fontsize)

fig2.subplots_adjust(left=0.092,bottom=0.13, right=0.95,top=0.895,wspace=0.22,hspace=0.316)
fig2.show()
filename = 'Fig/E2/'+'Net5Net0 pattern compare.pdf'
fig2.savefig(filename)

## net3  net 5
fig3 = plt.figure(figsize=(8,3))
fontsize = 8
a = Num.transpose(W_5_0_0)[:,10]
b = Num.transpose(W_3_0_1)[:,122-100]

axe = fig3.add_subplot(1, 1, 1)
axe.plot(a, label='Pattern1 in net5')
axe.plot(b, label='Pattern1 in net3')
axe.set_title('Pattern1 in net5 and net3', fontsize=fontsize)
axe.set_ylim([-4,6])
axe.legend(loc='upper right', fontsize=8)
axe.set_xlabel('Neuron index',fontsize=fontsize)
axe.set_ylabel('Pre-weights',fontsize=fontsize)

fig3.subplots_adjust(left=0.09, bottom=0.145, right=0.957,top=0.895,wspace=0.22,hspace=0.316)
fig3.show()
filename = 'Fig/E2/'+'Net5Net3 pattern compare.pdf'
fig3.savefig(filename)


a = 10
b = 10




## corcoe net5 and net0
# ar = np.reshape(a,(300,1))
# br = np.reshape(b,(300,1))
# cr = np.reshape(c,(300,1))
# dr = np.reshape(d,(300,1))
#
# df = np.concatenate((ar,br,cr,dr),axis=1)
# CorMat = np.corrcoef(df,rowvar=False)

# axe.plot(f, label='Neuron 163',color='#8c564b')




