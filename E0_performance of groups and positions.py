import numpy as np
import snntorch as snn
import torch
import numpy as Num
import math as Mat
import torch.nn as nn
from Fun.Fun_Enc import Fun_Enc
from Fun.Fun_Nor import Fun_Nor
from Fun.Fun_DatSeg import Fun_DatSeg
from torch.utils.data import TensorDataset, DataLoader
from torch.optim.lr_scheduler import ExponentialLR
from matplotlib import pyplot as plt
from DatPro import DatPro
from E0_Net import Net
from snntorch import spikegen
from nptdms import TdmsFile
from scipy.fftpack import fft
from Fun.DatSeg import DatSeg

keywords = 'd_1000-3000' # s_1000 s_2000 s_3000 d_1000-3000
DatSet = DatPro(Sec=10, Rat_Tra=0.8, keywords=keywords)

Tra = []
Tes = []
batch_size = 128
for i in range(6):
    Tra.append(DataLoader(DatSet[i*2], batch_size=batch_size, shuffle=True, drop_last=True))
    Tes.append(DataLoader(DatSet[i*2+1], batch_size=batch_size, shuffle=True, drop_last=True))
a=10
# plt.figure()
# plt.plot(DatSet[8][6500][0].squeeze())
# plt.plot(DatSet[8][2][0][0,:].squeeze()) double
# plt.plot(DatSet[8][2][0][1,:].squeeze()) double
# plt.plot(DatSet[8][0][0].squeeze())

#
# DatSet[4][0][1]  # data set  batch x/y
# DatSet[5][0][1]


a = 10


num_outputs = 2
num_steps = 100

loss_hist = []
test_loss_hist = []
counter = 0
beta = 0.95
data_type = torch.float32



def print_batch_accuracy(data, targets, train=False):
    output, _ = net(data.view(batch_size, -1))
    _, idx = output.sum(dim=0).max(1)
    acc = Num.mean((targets == idx).detach().cpu().numpy())

    if train:
        print(f"Train set accuracy for a single minibatch: {acc*100:.2f}%")
    else:
        print(f"Test set accuracy for a single minibatch: {acc*100:.2f}%")
def train_printer():
    print(f"Epoch {epoch}, Iteration {iter_counter}")
    print(f"Train Set Loss: {loss_hist[counter]:.2f}")
    print(f"Test Set Loss: {test_loss_hist[counter]:.2f}")
    print_batch_accuracy(data, targets, train=True)
    print_batch_accuracy(test_data, test_targets, train=False)
    # print("\n")
# Outer training loop


Num_Rep = 10 # 5
num_epochs = 2 # 10
Num_Dat = 6 #15
Ave = Num.zeros([Num_Rep, Num_Dat]) # j
Los = Num.zeros([Num_Rep, Num_Dat])

for j in range(3,Num_Dat):
    if j >= 3:
        FlaSen = 2
    else:
        FlaSen = 1
    for i in range(Num_Rep):
        num_inputs = DatSet[j*2][0][0].shape[1] * DatSet[j*2][0][0].shape[0]
        # num_hidden = int(200 / DatSet[j*2][0][0].shape[0])
        num_hidden = int(200)
        device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
        net = Net(num_inputs,num_hidden,num_outputs,beta,num_steps,FlaSen).to(device)
        loss = nn.CrossEntropyLoss()
        optimizer = torch.optim.Adam(net.parameters(), lr=0.01, betas=(0.9, 0.999))
        scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=3, gamma=0.1)
        for epoch in range(num_epochs):
            iter_counter = 0
            train_batch = iter(Tra[j])
            # scheduler.step()
            # Minibatch training loop
            for data, targets in train_batch:
                data = data.to(device)
                targets = targets.to(device)
                # a = data.numpy()[0,0,:,0]
                # b = data.numpy()[0,1,:,0]
                # c = data.view(128,-1).numpy()[0,:]
                # plt.plot(a)
                # plt.plot(b)
                # plt.plot(c)
                # forward pass
                net.train()
                spk_rec, mem_rec = net(data.view(batch_size, -1))

                # initialize the loss & sum over time
                loss_val = torch.zeros((1), dtype=data_type, device=device)
                for step in range(num_steps):
                    loss_val += loss(mem_rec[step], targets)

                # Gradient calculation + weight update
                optimizer.zero_grad()
                loss_val.backward()
                optimizer.step()

                # Store loss history for future plotting
                loss_hist.append(loss_val.item())

                # Test set
                with torch.no_grad():
                    net.eval()
                    test_data, test_targets = next(iter(Tes[j]))
                    test_data = test_data.to(device)
                    test_targets = test_targets.to(device)

                    # Test set forward pass
                    test_spk, test_mem = net(test_data.view(batch_size, -1))

                    # Test set loss
                    test_loss = torch.zeros((1), dtype=data_type, device=device)
                    for step in range(num_steps):
                        test_loss += loss(test_mem[step], test_targets)
                    test_loss_hist.append(test_loss.item())

                    # Print train/test loss/accuracy
                    if counter % 50 == 0:
                        train_printer()
                    counter += 1
                    iter_counter +=1
        # fig = plt.figure(facecolor="w", figsize=(10, 5))
        # plt.plot(loss_hist)
        # plt.plot(test_loss_hist)
        # plt.title("Loss Curves")
        # plt.legend(["Train Loss", "Test Loss"])
        # plt.xlabel("Iteration")
        # plt.ylabel("Loss")
        # plt.show()

        # for bat in range(0, batch_size):
        #     plt.subplot(1, 15, bat+1)
        #     plt.plot(test_spk[:, bat, 0])
        #     plt.plot(-1*test_spk[:, bat, 1])
        total = 0
        correct = 0
        test_loss_0 = []
        for data, targets in Tes[j]:
            with torch.no_grad():
                net.eval()
                data = data.to(device)
                targets = targets.to(device)

                # forward pass
                test_spk, test_mem = net(data.view(batch_size, -1))
                test_loss = torch.zeros((1), dtype=data_type, device=device)
                for step in range(num_steps):
                    test_loss += loss(test_mem[step], targets)
                test_loss_0.append(test_loss.item())

                # output, _ = net(data.view(batch_size, -1))
                # _, idx = output.sum(dim=0).max(1)
                # acc = Num.mean((targets == idx).detach().cpu().numpy())

                # calculate total accuracy
                _, predicted = test_spk.sum(dim=0).max(1)
                total += targets.size(0)
                correct += (predicted == targets).sum().item()
        print(f"Total correctly classified test set images: {correct}/{total}")
        print(f"Test Set Accuracy: {100 * correct / total:.2f}%")
        print('loss is', np.mean(np.array(test_loss_0)))
        Ave[i,j] = correct/total
        Los[i,j] = np.mean(np.array(test_loss_0))

        NetNam = 'Net/E0/'+keywords+f'/Net_{j:d}_{i:d}.pkl'
        torch.save(net, NetNam)

a=10
b=12

import openpyxl
data = Los
wb = openpyxl.Workbook()
ws_Los = wb.create_sheet('Los')
ws_Ave = wb.create_sheet('Acc')
for i in range(Los.shape[0]):  # Los.shape[0]
    for j in range(Los.shape[1]):  # Los.shape[1]
        ws_Los.cell(row=i+1, column=j+1, value=Los[i][j])
        ws_Ave.cell(row=i+1, column=j+1, value=Ave[i][j])
wb.save('Table/E0/'+keywords+'.xls')  ## table e0

a=10

